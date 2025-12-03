#!/usr/bin/env python
# -*- coding: utf-8 -*-
'''
@Time    : 2024/5/7 11:07 
@Author  : 
@File    : check_excel_data2024050702.py
@ProjectName: CRM 
'''
import datetime
import os
from decimal import Decimal

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import chardet

date_num = datetime.datetime.now().strftime("%Y%m%d%H%M")


def check_encoding(file_path):
    """
    检查文件编码是否为UTF-8
    :param file_path: 文件路径
    :return: True/False
    """
    with open(file_path, 'rb') as file:
        result = chardet.detect(file.read())
    if result['encoding'] == 'GB18030':
        return True
    else:
        return False


def detect_encoding(file_path):
    """
    检测文件编码
    :param file_path: 文件路径
    :return: 文件编码
    """
    with open(file_path, 'rb') as file:
        # result = chardet.detect(file.read(1024))  # 只读取文件的前1024字节
        result = chardet.detect(file.read())
    print(file_path, "检测到的文件编码为：", result['encoding'])
    if result['encoding'] == 'GB2312':
        return 'GB18030'
    else:
        return result['encoding']


def is_number(s):
    """
    判断字符串是否为数字
    :param s: 字符串
    :return: True/False
    """
    try:
        float(s)  # 尝试转换为浮点数，这会接受整数、浮点数，甚至是科学计数法
        return True
    except ValueError:
        return False


# def sort_data(file_path1, file_path2,):
#     """
#     对两个CSV文件进行排序
#     :param file_path1: 第一个CSV文件路径
#     :param file_path2: 第二个CSV文件路径
#     :return: None
#     """
#     df1 = pd.read_csv(file_path1, encoding=detect_encoding(file_path1)) # 读取csv文件，并指定编码格式为GB18030
#     df2 = pd.read_csv(file_path2, encoding=detect_encoding(file_path2))
#
#     df1.fillna('', inplace=True)  # 将空值替换为''
#     df2.fillna('', inplace=True)
#
#     # 筛选出包含'总计'的行
#     # df1_total = df1[df1['组别'] == '总计']
#     df1_total = df1[df1[df1.columns[0]] == '总计']
#     df2_total = df2[df2[df1.columns[0]] == '总计']
#
#     # 删除包含'总计'的行，以便进行排序
#     df1 = df1[df1[df1.columns[0]] != '总计']
#     df2 = df2[df2[df1.columns[0]] != '总计']
#
#     # df1 = df1.sort_values(by=df1.columns[1], ascending=True) # 按第二列进行升序排序
#     # df2 = df2.sort_values(by=df2.columns[1], ascending=True)
#     if '前端' in file_path1:
#         df1 = df1.sort_values(by=['一签率', '姓名'], ascending=[False, True])
#         df2 = df2.sort_values(by=['一签率', '姓名'], ascending=[False, True])
#     elif '复购' in file_path1:
#         df1 = df1.sort_values(by=['二签率', '姓名'], ascending=[False, True])
#         df2 = df2.sort_values(by=['二签率', '姓名'], ascending=[False, True])
#     else:
#         df1 = df1.sort_values(by=['三签率', '姓名'], ascending=[False, True])
#         df2 = df2.sort_values(by=['三签率', '姓名'], ascending=[False, True])
#     # 将排序后的数据与'总计'行重新组合
#     df1_final = pd.concat([df1, df1_total])
#     df2_final = pd.concat([df2, df2_total])
#     df1_final.to_csv(file_path1, index=False) # 保存到csv文件
#     df2_final.to_csv(file_path2, index=False)


def check_data(file_path1, file_path2, output_dir):
    """"
    检查两个CSV文件是否一致
    :param file_path1: 第一个CSV文件路径
    :param file_path2: 第二个CSV文件路径
    :param output_dir: 输出目录
    :return: None
    """
    # sort_data(file_path1, file_path2)

    df1 = pd.read_csv(file_path1, encoding=detect_encoding(file_path1))
    df2 = pd.read_csv(file_path2, encoding=detect_encoding(file_path2))

    df1.fillna('', inplace=True)  # 将空值替换为''
    df2.fillna('', inplace=True)


    # df1 = pd.read_csv(file_path1)
    # # df1 = pd.read_csv(file_path1, encoding='GB18030')  # 读取csv文件，并指定编码格式为GB18030
    # # df1 = df1.applymap(lambda x: x.encode('utf-8').decode('utf-8') if isinstance(x, str) else x) # 将中文字符转换为utf-8编码
    # df1.fillna('', inplace=True)  # 将空值替换为''
    # df2 = pd.read_csv(file_path2)
    # # df2 = pd.read_csv(file_path2, encoding='GB18030')  # 读取csv文件，并指定编码格式为GB18030
    # df2.fillna('', inplace=True)

    all_data = pd.concat([df1, df2], ignore_index=False, axis=1)
    all_data.to_excel(os.path.join(output_dir, '{}{}-{}.xlsx'.format(file_results, date_num, i)), index=True,
                          header=True)
    # 输出all_data及文字说明
    print("合并数据结果：\n{}".format(all_data))

    highlight_differences(file_path_diff)


def highlight_differences(output_dir):
    """"
    高亮差异数据
    :param output_dir: 输出目录
    :return: None
    """
    # 加载已存在的Excel文件
    wb = load_workbook(filename=os.path.join(output_dir, '{}{}-{}.xlsx'.format(file_results, date_num, i)))
    ws = wb[wb.sheetnames[0]]  # 要处理的是第一个Sheet的工作表
    # ws = wb['差异数据']   # 处理的是名为'差异数据'的工作表
    num_columns = ws.max_column

    # 填充的高亮颜色
    # 红色高亮颜色
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    # 黄色高亮颜色
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    # 绿色高亮颜色
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    # 遍历工作表的行和对比名称相同的列
    for row in ws.iter_rows(min_row=2):  # 第二行开始避免标题行
        for col_index in range(0, num_columns // 2):
            # print(type(row[col_index + 1].value), type(row[num_columns // 2 + col_index + 1].value))
            # print(row[col_index + 1].value, row[num_columns // 2 + col_index + 1].value)
            if str(row[col_index + 1].value) != str(row[num_columns // 2 + col_index + 1].value):
                row[col_index + 1].fill = red_fill
                row[num_columns // 2 + col_index + 1].fill = red_fill
                # print(f"{row[col_index + 1].value} 和 {row[num_columns // 2 + col_index + 1].value} 不相等")
                if row[col_index + 1].value is not None and row[num_columns // 2 + col_index + 1].value is not None:
                    row[col_index + 1].value = str(row[col_index + 1].value)
                    row[num_columns // 2 + col_index + 1].value = str(row[num_columns // 2 + col_index + 1].value)
                    # 判断是否为百分比
                    if "%" in row[col_index + 1].value and "%" in row[num_columns // 2 + col_index + 1].value:
                        # print(f"{row[col_index + 1].value} 和 {row[num_columns // 2 + col_index + 1].value} 都为百分比")
                        row[col_index + 1].value = row[col_index + 1].value.strip('%')  # 去除百分号
                        row[num_columns // 2 + col_index + 1].value = row[num_columns // 2 + col_index + 1].value.strip(
                            '%')
                        # print(f"{row[col_index + 1].value} 和 {row[num_columns // 2 + col_index + 1].value} 去除百分号")
                        # 判断绝对值
                        if abs(Decimal(float(row[col_index + 1].value)) - Decimal(
                                float(row[num_columns // 2 + col_index + 1].value))) < 0.011:
                            print(f"百分比：{row[col_index + 1].value} - {row[num_columns // 2 + col_index + 1].value} 绝对值= {abs(Decimal(float(row[col_index + 1].value)) - Decimal(float(row[num_columns // 2 + col_index + 1].value)))}")
                            row[col_index + 1].fill = yellow_fill
                            row[num_columns // 2 + col_index + 1].fill = yellow_fill
                        # elif abs(Decimal(float(row[col_index + 1].value)) - Decimal(
                        #         float(row[num_columns // 2 + col_index + 1].value))) > 0.011:
                        #     row[col_index + 1].fill = red_fill
                        #     row[num_columns // 2 + col_index + 1].fill = red_fill
                        # 添加百分号
                        row[col_index + 1].value = str(row[col_index + 1].value) + '%'
                        row[num_columns // 2 + col_index + 1].value = str(
                            row[num_columns // 2 + col_index + 1].value) + '%'
                    # 判断是否为数字
                    # elif row[col_index + 1].value.isdigit() and row[num_columns // 2 + col_index + 1].value.isdigit() or row[col_index + 1].value.strip('.') and row[num_columns // 2 + col_index + 1].value.strip('.'):
                    elif is_number(row[col_index + 1].value) and is_number(row[num_columns // 2 + col_index + 1].value):
                        if abs(eval(row[col_index + 1].value) - eval(
                                row[num_columns // 2 + col_index + 1].value)) < 0.011:  # 判断绝对值
                            # if abs(Decimal(float(row[col_index + 1].value)) - Decimal(float(row[num_columns // 2 + col_index + 1].value))) < 0.011:
                            print(f"数字：{row[col_index + 1].value} - {row[num_columns // 2 + col_index + 1].value} 绝对值= {abs(eval(row[col_index + 1].value) - eval(row[num_columns // 2 + col_index + 1].value))}")
                            row[col_index + 1].fill = yellow_fill
                            row[num_columns // 2 + col_index + 1].fill = yellow_fill
                        # elif abs(eval(row[col_index + 1].value) - eval(row[num_columns // 2 + col_index + 1].value)) >= 0.011:
                        #     row[col_index + 1].fill = red_fill
                        #     row[num_columns // 2 + col_index + 1].fill = red_fill
            # else:
            #     row[col_index + 1].fill = green_fill
            #     row[num_columns // 2 + col_index + 1].fill = green_fill

    # 保存更改后的Excel文件
    wb.close()
    wb.save(os.path.join(output_dir, '{}{}-{}.xlsx'.format(file_results, date_num, i)))

# old_data = ['02-05生产8月-ROI统计_按渠道-百度-信息流-免费票-百度推广_总ROI202404021659.csv',
#            '02-05生产8月-ROI统计_按渠道-百度-信息流-免费票-百度推广_前端ROI202404021659.csv',
#            '02-05生产8月-ROI统计_按渠道-百度-信息流-免费票-百度推广_复购ROI202404021659.csv',
#            '02-05生产8月-ROI统计_按渠道-百度-信息流-免费票-百度推广_客服ROI202404021659.csv']
# new_data = ['02-05生产8月-ROI统计Task_按渠道-百度-信息流-免费票-百度推广_总ROI202404021659.csv',
#             '02-05生产8月-ROI统计Task_按渠道-百度-信息流-免费票-百度推广_前端ROI202404021659.csv',
#             '02-05生产8月-ROI统计Task_按渠道-百度-信息流-免费票-百度推广_复购ROI202404021700.csv',
#             '02-05生产8月-ROI统计Task_按渠道-百度-信息流-免费票-百度推广_客服ROI202404021700.csv']


old_data = ['推广_进线统计极速版数据20250801~20250810_按渠道-字节-亚商抖音直播_zbsk5-财经潇潇子.csv']
new_data = ['推广_进线统计生产数据20250801~20250810_按渠道-字节-亚商抖音直播_zbsk5-财经潇潇子.csv']

# old_data = ['推广_团队数据-客服-202403按公司.csv']
# new_data = ['推广_团队数据极速版-客服-202403按公司.csv']


file_path_source = r'D:\亚商'
file_path_diff = r'D:\亚商\数据核对'
file_results = '进线统计+极速版-20250801~20250810_按渠道-字节-亚商抖音直播_zbsk5-财经潇潇子_数据核对结果'
# _营销中心-一中心-二区-二部-一组_是自然_否二维码

i = 0
for old, new in zip(old_data, new_data):
    i = i + 1
    check_data('{}\\{}'.format(file_path_source, old), '{}\\{}'.format(file_path_source, new), file_path_diff)
