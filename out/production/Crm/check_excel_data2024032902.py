#!/usr/bin/env python
# -*- coding: utf-8 -*-
'''
@Time    : 2024/3/29 16:37 
@Author  : admin
@File    : check_excel_data2024032902.py
@ProjectName: CRM 
'''
import os
import pandas as pd
import datetime

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

date = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
date_num = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
# print(date, date_num)

# def delete_files(folder_path):
#     for filename in os.listdir(folder_path):
#         file_path = os.path.join(folder_path, filename)
#
#         # 判断是否为文件
#         if os.path.isfile(file_path):
#             os.remove(file_path)
#         # 如果是子文件夹，递归删除其中的文件
#         elif os.path.isdir(file_path):
#             delete_files(file_path)
#         print(f"{date} 删除文件 {file_path}")


def check_data(file_path1, file_path2, output_dir):
    df1 = pd.read_csv(file_path1, encoding='gb18030')
    df2 = pd.read_csv(file_path2, encoding='gb18030')

    # 找到两个文件中完全相同的行
    same_rows = df1.merge(df2, indicator=True, how='outer').query('_merge == "both"').drop('_merge', axis=1)


    # 找到在df1中但不在df2中的行
    in_df1_not_df2 = df1[~df1.index.isin(same_rows.index)].dropna(subset=df1.columns)

    # 找到在df2中但不在df1中的行
    in_df2_not_df1 = df2[~df2.index.isin(same_rows.index)].dropna(subset=df2.columns)

    # 按列合并（axis=1）所有差异行的DataFrame
    diff_results = pd.concat([in_df1_not_df2, in_df2_not_df1], ignore_index=False, axis=1)
    print("行数：{}，列数：{}".format(diff_results.shape[0], diff_results.shape[1]))
    print(len(diff_results))


    with pd.ExcelWriter(os.path.join(output_dir, '{}{}.xlsx'.format(file_name, date_num))) as writer:
        # i = 0
        for sheet in same_rows, diff_results:
            # i = i+1
            # sheet.to_excel(writer, sheet_name='核对结果{}'.format(i))
            if sheet is same_rows:
                sheet.to_excel(writer, sheet_name='相同行')
            # elif sheet is in_df1_not_df2:
            #     sheet.to_excel(writer, sheet_name='df1中特有的行')
            else:
                sheet.to_excel(writer, sheet_name='差异数据')

        print(f"数据核对结果已输出到 {os.path.join(output_dir, '{}{}.xlsx'.format(file_name, date_num))}")


    # 加载已存在的Excel文件
    wb = load_workbook(filename=os.path.join(output_dir, '{}{}.xlsx'.format(file_name, date_num)))
    ws = wb[wb.sheetnames[1]]  # 要处理的是第二个Sheet的工作表
    num_columns = ws.max_column

    # 填充的高亮颜色
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 黄色高亮

    # 遍历工作表的行和对比名称相同的列
    for row in ws.iter_rows(min_row=2):  #
        print("row：{}".format(row))
        for col_index in range(0, num_columns // 2):
            print("col_index：{}".format(col_index))
            if row[col_index + 1].value != row[num_columns // 2 + col_index + 1].value:
                row[col_index + 1].fill = highlight_fill
                row[num_columns // 2 + col_index + 1].fill = highlight_fill

    # 保存更改后的Excel文件
    wb.close()
    wb.save(os.path.join(output_dir, '{}{}.xlsx'.format(file_name, date_num)))


# 使用方法
# delete_files(r'D:\亚商\数据核对')
file_name = '数据核对结果'
check_data(r'D:\亚商\ROI统计_总ROI_2023-03_按公司.csv', r'D:\亚商\ROI统计Task_总ROI_2023-03_按公司.csv', r'D:\亚商\数据核对')