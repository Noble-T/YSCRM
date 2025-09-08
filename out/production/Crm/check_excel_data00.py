#!/usr/bin/env python
# -*- coding: utf-8 -*-
'''
@Time    : 2024/4/1 15:06 
@Author  : 
@File    : check_excel_data00.py
@ProjectName: CRM 
'''
import os
from decimal import Decimal

import pandas as pd
import datetime

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

date = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
date_num = datetime.datetime.now().strftime("%Y%m%d%H%M")
# print(date, date_num)

# def delete_files(folder_path):
#     for filename in os.listdir(folder_path):
#         file_path = os.path.join(folder_path, filename)
#
#         # 判断是否为文件
#         if os.path.isfile(file_path):
#             os.remove(file_path)
#         # 如果是子文件夹，递归删除其中的文件
#         elif os.path.isdir(file_path):
#             delete_files(file_path)
#         print(f"{date} 删除文件 {file_path}")



def is_number(s):
    try:
        float(s)  # 尝试转换为浮点数，这会接受整数、浮点数，甚至是科学计数法
        return True
    except ValueError:
        return False

def check_data(file_path1, file_path2, output_dir):
    # df1 = pd.read_csv(file_path1, encoding='GB18030') # 读取csv文件，并指定编码格式为GB18030
    df1 = pd.read_csv(file_path1)
    # df1 = df1.applymap(lambda x: x.encode('utf-8').decode('utf-8') if isinstance(x, str) else x) # 将中文字符转换为utf-8编码
    df1.fillna('', inplace=True)     # 将空值替换为''
    df2 = pd.read_csv(file_path2)
    df2.fillna('', inplace=True)

    # 删除重复的行
    # diff = pd.concat([df1, df2]).drop_duplicates(keep=False)
    # 重复的行去重
    # diff = pd.merge(df1, df2, on=df1.columns.tolist(), how='inner')
    diff = pd.concat([df1, df2], axis=0, ignore_index=True)
    diff = diff.drop_duplicates(keep=False)
    # same_data = pd.merge(df1, df2, how='inner', on=None, suffixes=('_df1', '_df2'))
    print(f"\n\n==========差异数据：\n{diff}")
    diff.to_excel(os.path.join(output_dir, 'diff{}.xlsx'.format(date_num)), index=True, header=True)
    # print(f"差异数据文件 {os.path.join(output_dir, 'diff{}-{}.xlsx'.format(date_num, i))}")

    # 找到在df1中但不在diff中的行
    same_diff = df1[~df1.index.isin(diff.index)].dropna(subset=df1.columns)
    print(f"相同数据：\n{same_diff}")
    same_diff.to_excel(os.path.join(output_dir, 'same_diff{}.xlsx'.format(date_num)), header=True)


    # 找到在df1中但不在same_diff中的行
    in_df1_not_same = df1[~df1.index.isin(same_diff.index)].dropna(subset=df1.columns)
    print(f"老表差异数据：\n{in_df1_not_same}")

    # 找到在df2中但不在same_diff中的行
    in_df2_not_same = df2[~df2.index.isin(same_diff.index)].dropna(subset=df2.columns)
    print(f"新表差异数据：\n{in_df2_not_same}")

    # 按列合并（axis=1）所有差异行的DataFrame
    diff_results = pd.concat([in_df1_not_same, in_df2_not_same], ignore_index=False, axis=1)
    print(f"合并差异数据：\n{diff_results}")

    with pd.ExcelWriter(os.path.join(output_dir, '{}{}-{}.xlsx'.format(file_results, date_num, i))) as writer:
        df1.to_excel(writer, sheet_name='old')
        df2.to_excel(writer, sheet_name='new')
        for sheet in same_diff, diff_results:
            # sheet.to_excel(writer, sheet_name='核对结果{}'.format(i))
            if sheet is same_diff:
                sheet.to_excel(writer, sheet_name='相同数据')
            else:
                sheet.to_excel(writer, sheet_name='差异数据')

        print(f"数据核对结果已输出到 {os.path.join(output_dir, '{}{}-{}.xlsx'.format(file_results, date_num, i))}")

    highlight_differences(file_path_diff)


def highlight_differences(output_dir):
    # 加载已存在的Excel文件
    wb = load_workbook(filename=os.path.join(output_dir, '{}{}-{}.xlsx'.format(file_results, date_num, i)))
    # ws = wb[wb.sheetnames[1]]  # 要处理的是第二个Sheet的工作表
    ws = wb['差异数据']   # 处理的是名为'差异数据'的工作表
    num_columns = ws.max_column

    # 填充的高亮颜色
    # 黄色高亮颜色
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    # 绿色高亮颜色
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")


    # 遍历工作表的行和对比名称相同的列
    for row in ws.iter_rows(min_row=2):  # 第二行开始避免标题行
        for col_index in range(0, num_columns // 2):
            # print(type(row[col_index + 1].value), type(row[num_columns // 2 + col_index + 1].value))
            if row[col_index + 1].value != row[num_columns // 2 + col_index + 1].value:
                row[col_index + 1].fill = yellow_fill
                row[num_columns // 2 + col_index + 1].fill = yellow_fill
                # print(f"{row[col_index + 1].value} 和 {row[num_columns // 2 + col_index + 1].value} 不相等")
                if row[col_index + 1].value is not None and row[num_columns // 2 + col_index + 1].value is not None:
                    row[col_index + 1].value = str(row[col_index + 1].value)
                    row[num_columns // 2 + col_index + 1].value = str(row[num_columns // 2 + col_index + 1].value)
                    # 判断是否为百分比
                    if "%" in row[col_index + 1].value and "%" in row[num_columns // 2 + col_index + 1].value:
                        # print(f"{row[col_index + 1].value} 和 {row[num_columns // 2 + col_index + 1].value} 都为百分比")
                        row[col_index + 1].value = row[col_index + 1].value.strip('%')  # 去除百分号
                        row[num_columns // 2 + col_index + 1].value = row[num_columns // 2 + col_index + 1].value.strip('%')
                        # print(f"{row[col_index + 1].value} 和 {row[num_columns // 2 + col_index + 1].value} 去除百分号")
                        # 判断绝对值
                        if abs(Decimal(float(row[col_index + 1].value)) - Decimal(float(row[num_columns // 2 + col_index + 1].value))) < 0.011:
                            # print(f"百分比：{row[col_index + 1].value} - {row[num_columns // 2 + col_index + 1].value} 绝对值= {abs(Decimal(float(row[col_index + 1].value)) - Decimal(float(row[num_columns // 2 + col_index + 1].value)))}")
                            row[col_index + 1].fill = green_fill
                            row[num_columns // 2 + col_index + 1].fill = green_fill
                        # 添加百分号
                        row[col_index + 1].value = str(row[col_index + 1].value) + '%'
                        row[num_columns // 2 + col_index + 1].value = str(row[num_columns // 2 + col_index + 1].value) + '%'
                    # 判断是否为数字
                    # elif row[col_index + 1].value.isdigit() and row[num_columns // 2 + col_index + 1].value.isdigit() or row[col_index + 1].value.strip('.') and row[num_columns // 2 + col_index + 1].value.strip('.'):
                    elif is_number(row[col_index + 1].value) and is_number(row[num_columns // 2 + col_index + 1].value):
                        if abs(eval(row[col_index + 1].value) - eval(row[num_columns // 2 + col_index + 1].value)) < 0.011:  # 判断绝对值
                        # if abs(Decimal(float(row[col_index + 1].value)) - Decimal(float(row[num_columns // 2 + col_index + 1].value))) < 0.011:
                            # print(f"数字：{row[col_index + 1].value} - {row[num_columns // 2 + col_index + 1].value} 绝对值= {abs(eval(row[col_index + 1].value) - eval(row[num_columns // 2 + col_index + 1].value))}")
                            row[col_index + 1].fill = green_fill
                            row[num_columns // 2 + col_index + 1].fill = green_fill


    # 保存更改后的Excel文件
    wb.close()
    wb.save(os.path.join(output_dir, '{}{}-{}.xlsx'.format(file_results, date_num, i)))


# 使用方法
# delete_files(r'D:\亚商\数据核对')
# check_data(r'D:\亚商\生产_ROI统计_按公司_总ROI202404011705.csv', r'D:\亚商\生产_ROI统计Task_按公司_总ROI202404011705.csv', r'D:\亚商\数据核对')


# task_old = ['02-05生产8月-ROI统计_按渠道-百度-信息流-免费票-百度推广_总ROI202404021659.csv',
#            '02-05生产8月-ROI统计_按渠道-百度-信息流-免费票-百度推广_前端ROI202404021659.csv',
#            '02-05生产8月-ROI统计_按渠道-百度-信息流-免费票-百度推广_复购ROI202404021659.csv',
#            '02-05生产8月-ROI统计_按渠道-百度-信息流-免费票-百度推广_客服ROI202404021659.csv']
# task_new = ['02-05生产8月-ROI统计Task_按渠道-百度-信息流-免费票-百度推广_总ROI202404021659.csv',
#             '02-05生产8月-ROI统计Task_按渠道-百度-信息流-免费票-百度推广_前端ROI202404021659.csv',
#             '02-05生产8月-ROI统计Task_按渠道-百度-信息流-免费票-百度推广_复购ROI202404021700.csv',
#             '02-05生产8月-ROI统计Task_按渠道-百度-信息流-免费票-百度推广_客服ROI202404021700.csv']


# old_data = ['推广统计数据-生产202309-按公司_亚商_百度_百度信息流_百度免费票_百度推广-202404121347.csv']
# new_data = ['推广统计数据Task-生产202309-按公司_亚商_百度_百度信息流_百度免费票_百度推广-202404121347.csv']

# old_data = ['推广_员工数据-前端-202403按公司_亚商_百度_信息流_百度免费票_百度推广.csv']
# new_data = ['推广_员工数据极速版-前端-202403按公司_亚商_百度_信息流_百度免费票_百度推广.csv']

# old_data = ['推广_团队数据-前端-202403按公司.csv']
# new_data = ['推广_团队数据极速版-前端-202403按公司.csv']

old_data = ['推广_员工数据-复购-202403按公司.csv']
new_data = ['推广_员工数据极速版-复购-202403按公司.csv']

# old_data = ['推广_团队数据-客服-202403按公司.csv']
# new_data = ['推广_团队数据极速版-客服-202403按公司.csv']


file_path_source = r'D:\亚商'
file_path_diff = r'D:\亚商\数据核对'
file_results = '推广_员工数据-复购-202403按公司-数据核对结果'

i = 0
for old, new in zip(old_data, new_data):
    i = i + 1
    check_data('{}\\{}'.format(file_path_source, old), '{}\\{}'.format(file_path_source, new), file_path_diff)

